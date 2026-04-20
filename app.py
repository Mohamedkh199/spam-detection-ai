import json
import os
import pickle
from flask import Flask, render_template, request, redirect
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint

app = Flask(__name__)

history_data = []
EXCEL_FILE = "history.xlsx"

# تحميل الموديل والـ vectorizer
with open("model.pkl", "rb") as f:
    model = pickle.load(f)

with open("vectorizer.pkl", "rb") as f:
    vectorizer = pickle.load(f)

def save_to_excel(message, prediction, time):
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "History"
        ws.append(["Message", "Prediction", "Time"])

    ws.append([message, prediction, time])

    # احسب التوزيع
    spam_count = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if row[1] == "Spam ❌")
    notspam_count = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if row[1] == "Not Spam ✅")
    neutral_count = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if row[1] == "Neutral 🔵")

    ws["E1"] = "Category"
    ws["F1"] = "Count"
    ws["E2"] = "Spam"
    ws["F2"] = spam_count
    ws["E3"] = "Not Spam"
    ws["F3"] = notspam_count
    ws["E4"] = "Neutral"
    ws["F4"] = neutral_count

    # أنشئ الرسم
    pie = PieChart()
    labels = Reference(ws, min_col=5, min_row=2, max_row=4)
    data = Reference(ws, min_col=6, min_row=1, max_row=4)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Distribution of Checked Messages"

    # ضبط الألوان: Spam أحمر، Not Spam أخضر، Neutral أزرق
    series = pie.series[0]
    series.points = []

    dp1 = DataPoint(idx=0)
    dp1.graphicalProperties.solidFill = "FF0000"  # Spam = أحمر
    series.points.append(dp1)

    dp2 = DataPoint(idx=1)
    dp2.graphicalProperties.solidFill = "00FF00"  # Not Spam = أخضر
    series.points.append(dp2)

    dp3 = DataPoint(idx=2)
    dp3.graphicalProperties.solidFill = "0000FF"  # Neutral = أزرق
    series.points.append(dp3)

    ws.add_chart(pie, "H2")
    wb.save(EXCEL_FILE)

@app.route('/', methods=['GET', 'POST'])
def check():
    prediction = None
    message = request.form.get('message', '') or request.args.get('message', '')
    action = request.form.get('action', '')

    if action == "model_info":
        return redirect("/model")
    if action == "sample_message":
        return redirect("/samples")

    # الحالة: ضغط زر Check Message
    if request.method == 'POST' and action == "check":
        if message.strip() == "":   # إذا الصندوق فاضي
            prediction = "Neutral 🔵"
            history_data.append({
                "message": "(Empty Message)",
                "prediction": prediction,
                "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })
            save_to_excel("(Empty Message)", prediction, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        else:
            # استخدم الموديل للتنبؤ
            X = vectorizer.transform([message])
            y_pred = model.predict(X)[0]

            if y_pred == 1:   # 1 = Spam
                prediction = "Spam ❌"
            else:             # 0 = Not Spam
                prediction = "Not Spam ✅"

            history_data.append({
                "message": message,
                "prediction": prediction,
                "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })
            save_to_excel(message, prediction, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    return render_template('check.html', prediction=prediction, message=message)

@app.route('/history')
def history():
    spam_count = sum(1 for item in history_data if item["prediction"] == "Spam ❌")
    notspam_count = sum(1 for item in history_data if item["prediction"] == "Not Spam ✅")
    neutral_count = sum(1 for item in history_data if item["prediction"] == "Neutral 🔵")

    return render_template(
        'history.html',
        history=history_data,
        spam_count=spam_count,
        notspam_count=notspam_count,
        neutral_count=neutral_count
    )

@app.route('/reset')
def reset_history():
    global history_data
    history_data = []  # امسح البيانات من الذاكرة

    try:
        if os.path.exists(EXCEL_FILE):
            os.remove(EXCEL_FILE)
    except PermissionError:
        print("⚠️ الملف history.xlsx مفتوح في Excel، سكّريه أولاً عشان أقدر أمسحه.")

    return redirect('/history')

@app.route('/model')
def model_info():
    model_details = {
        "name": "Naive Bayes Classifier",
        "description": "This model uses TF-IDF Vectorizer to transform text data and Naive Bayes for classification.",
        "implementation": "Project implementation by: Mohammad Hayajneh and Mothana Mufleh",
        "supervised": "Supervised by: Dr. Ra’ed Alkhateeb"
    }
    return render_template('model.html', model=model_details)

@app.route('/samples')
def samples():
    try:
        with open("messages.json", "r", encoding="utf-8") as f:
            data = json.load(f)
        return render_template('samples.html', spam=data["spam"], not_spam=data["not_spam"])
    except FileNotFoundError:
        return "messages.json file not found. Please make sure it's in the same folder as app.py."




if __name__ == '__main__':
   
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port)